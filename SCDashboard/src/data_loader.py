from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from .config import (
    ASSESSMENTS_FILE,
    EVIDENCE_FILE,
    PROJECTS_FILE,
    RECOMMENDATIONS_FILE,
    STANDARDS_FILE,
)
from .scoring import enrich_assessments


def _clean_text(value: object) -> object:
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value).strip()
        return cleaned if cleaned else pd.NA
    return value


def _clean_frame(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        result[column] = result[column].map(_clean_text)
    return result


def _read_merged_positions(
    path: Path,
    sheet_name: str | int,
    column_positions: list[int],
    column_names: list[str],
) -> pd.DataFrame:
    """Read source workbooks that use horizontally merged cells."""
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
    if raw.shape[1] <= max(column_positions):
        raise ValueError(
            f"{path.name}/{sheet_name} has {raw.shape[1]} columns, but the dashboard "
            f"expects at least {max(column_positions) + 1}."
        )
    selected = raw.iloc[:, column_positions].copy()
    selected.columns = column_names
    return _clean_frame(selected)


def _read_hyperlink_targets(
    path: Path,
    sheet_name: str,
    key_column: int,
    link_column: int,
) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook[sheet_name]
        targets: dict[str, str] = {}
        for row in sheet.iter_rows():
            key = row[key_column - 1].value
            link_cell = row[link_column - 1]
            if key and link_cell.hyperlink and link_cell.hyperlink.target:
                targets[str(key).strip().casefold()] = link_cell.hyperlink.target
        return targets
    finally:
        workbook.close()


def load_standards(path: Path = STANDARDS_FILE) -> pd.DataFrame:
    standards = _read_merged_positions(
        path,
        "Sheet1",
        [0, 1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31, 34],
        [
            "indicator_id",
            "domain",
            "subdomain",
            "indicator_name",
            "indicator_definition",
            "required_data",
            "required_evidence",
            "project_type_tags",
            "standard_source",
            "measurement_method",
            "applicability",
            "limitations",
            "maturity_guidance",
        ],
    )
    standards["indicator_id"] = standards["indicator_id"].astype("string").str.strip()
    standards = standards[
        standards["indicator_id"].str.match(r"^[MEG]\d+$", na=False)
    ].copy()
    standards = standards.drop_duplicates("indicator_id").reset_index(drop=True)
    standards["standard_source"] = standards["standard_source"].fillna(
        "SCI Smart City Standards Project"
    )
    return standards


def load_projects(path: Path = PROJECTS_FILE) -> pd.DataFrame:
    projects = _read_merged_positions(
        path,
        "Sheet1",
        [0, 1, 4, 7, 10, 13, 16, 19, 22, 25],
        [
            "project_id",
            "project_name",
            "domain",
            "project_type",
            "description",
            "technology_used",
            "target_users",
            "data_sources",
            "current_status",
            "expected_impact",
        ],
    )
    projects["project_id"] = projects["project_id"].astype("string").str.strip()
    projects = projects[projects["project_id"].str.match(r"^P\d+$", na=False)].copy()
    return projects.drop_duplicates("project_id").reset_index(drop=True)


def _split_indicator_ids(value: object) -> list[str]:
    if value is None or pd.isna(value):
        return []
    return [part.strip() for part in re.split(r"[,;/]", str(value)) if part.strip()]


def load_project_indicator_mapping(path: Path = PROJECTS_FILE) -> pd.DataFrame:
    mapping = _read_merged_positions(
        path,
        "Sheet2",
        [0, 1, 2, 5],
        ["project_id", "indicator_id", "indicator_name_from_mapping", "why_it_applies"],
    )
    mapping["project_id"] = mapping["project_id"].astype("string").str.strip()
    mapping = mapping[mapping["project_id"].str.match(r"^P\d+$", na=False)].copy()
    mapping["indicator_id"] = mapping["indicator_id"].apply(_split_indicator_ids)
    mapping = mapping.explode("indicator_id", ignore_index=True)
    mapping["indicator_id"] = mapping["indicator_id"].astype("string").str.strip()
    mapping = mapping[
        mapping["indicator_id"].str.match(r"^[MEG]\d+$", na=False)
    ].copy()
    return mapping.drop_duplicates(["project_id", "indicator_id"]).reset_index(drop=True)


def _parse_excel_date(value: object) -> pd.Timestamp | pd.NaT:
    if value is None or pd.isna(value):
        return pd.NaT
    if isinstance(value, (int, float)):
        try:
            return pd.to_datetime(value, unit="D", origin="1899-12-30")
        except (ValueError, TypeError, OverflowError):
            return pd.NaT
    return pd.to_datetime(value, errors="coerce")


def load_evidence(path: Path = EVIDENCE_FILE) -> pd.DataFrame:
    evidence = _read_merged_positions(
        path,
        "Sheet1",
        [0, 1, 2, 3, 6, 9, 12, 15, 18, 21, 24, 27],
        [
            "evidence_id",
            "project_id",
            "indicator_id",
            "evidence_title",
            "evidence_type",
            "evidence_source",
            "source_organization",
            "evidence_date",
            "evidence_quality_score",
            "evidence_quality_level_source",
            "link_or_reference",
            "evidence_notes",
        ],
    )
    evidence["evidence_id"] = evidence["evidence_id"].astype("string").str.strip()
    evidence = evidence[
        evidence["evidence_id"].str.match(r"^Ev\d+$", case=False, na=False)
    ].copy()
    evidence["project_id"] = evidence["project_id"].astype("string").str.strip()
    evidence["indicator_id"] = evidence["indicator_id"].apply(_split_indicator_ids)
    evidence = evidence.explode("indicator_id", ignore_index=True)
    evidence["indicator_id"] = evidence["indicator_id"].astype("string").str.strip()
    evidence["evidence_quality_score"] = pd.to_numeric(
        evidence["evidence_quality_score"], errors="coerce"
    ).fillna(0).clip(0, 4)
    evidence["evidence_date"] = evidence["evidence_date"].map(_parse_excel_date)
    link_targets = _read_hyperlink_targets(path, "Sheet1", 1, 25)
    evidence["link_url"] = (
        evidence["evidence_id"].astype(str).str.casefold().map(link_targets)
    )
    return evidence.reset_index(drop=True)


def load_assessments(path: Path = ASSESSMENTS_FILE) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing {path.name} inside the data folder.")
    assessments = _clean_frame(pd.read_csv(path))
    required = {
        "assessment_id",
        "project_id",
        "indicator_id",
        "applicability_status",
        "current_score",
        "target_score",
        "evidence_quality_score",
        "data_readiness_score",
        "high_impact_gap",
        "assessment_notes",
    }
    missing = required.difference(assessments.columns)
    if missing:
        raise ValueError(f"assessment_scores.csv is missing columns: {sorted(missing)}")
    assessments["project_id"] = assessments["project_id"].astype("string").str.strip()
    assessments["indicator_id"] = assessments["indicator_id"].astype("string").str.strip()
    return assessments


def load_recommendations(path: Path = RECOMMENDATIONS_FILE) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(
            columns=[
                "recommendation_id",
                "project_id",
                "indicator_id",
                "gap_type",
                "recommended_action",
                "priority_level_source",
                "expected_benefit",
                "estimated_effort",
                "evidence_needed",
                "responsible_party_type",
            ]
        )
    recommendations = _clean_frame(pd.read_csv(path))
    required = {
        "recommendation_id",
        "project_id",
        "indicator_id",
        "recommended_action",
    }
    missing = required.difference(recommendations.columns)
    if missing:
        raise ValueError(f"recommendations.csv is missing columns: {sorted(missing)}")
    recommendations["project_id"] = recommendations["project_id"].astype("string").str.strip()
    recommendations["indicator_id"] = recommendations["indicator_id"].astype("string").str.strip()
    return recommendations.drop_duplicates(["project_id", "indicator_id"]).reset_index(drop=True)


def _join_unique(series: pd.Series) -> object:
    values: list[str] = []
    for value in series.dropna():
        text = str(value).strip()
        if text and text not in values:
            values.append(text)
    return " | ".join(values) if values else pd.NA


def _evidence_summary(evidence: pd.DataFrame) -> pd.DataFrame:
    if evidence.empty:
        return pd.DataFrame(columns=["project_id", "indicator_id"])
    return (
        evidence.groupby(["project_id", "indicator_id"], as_index=False)
        .agg(
            repository_evidence_quality_score=("evidence_quality_score", "max"),
            evidence_count=("evidence_id", "nunique"),
            evidence_id=("evidence_id", _join_unique),
            evidence_title=("evidence_title", _join_unique),
            evidence_type=("evidence_type", _join_unique),
            evidence_source=("evidence_source", _join_unique),
            source_organization=("source_organization", _join_unique),
            evidence_date=("evidence_date", "max"),
            link_or_reference=("link_or_reference", _join_unique),
            evidence_notes=("evidence_notes", _join_unique),
        )
    )


def merge_dashboard_data(
    assessments: pd.DataFrame,
    standards: pd.DataFrame,
    projects: pd.DataFrame,
    mapping: pd.DataFrame,
    evidence: pd.DataFrame,
    recommendations: pd.DataFrame,
) -> pd.DataFrame:
    merged = assessments.copy()
    merged = merged[merged["applicability_status"].fillna("").eq("Applicable")].copy()

    merged = merged.merge(projects, on="project_id", how="left", validate="many_to_one")
    merged = merged.merge(
        standards,
        on="indicator_id",
        how="left",
        validate="many_to_one",
        suffixes=("", "_standard"),
    )

    applicability = mapping[
        ["project_id", "indicator_id", "why_it_applies"]
    ].drop_duplicates(["project_id", "indicator_id"])
    merged = merged.merge(applicability, on=["project_id", "indicator_id"], how="left")
    merged = merged.merge(
        _evidence_summary(evidence),
        on=["project_id", "indicator_id"],
        how="left",
        validate="one_to_one",
    )
    merged = merged.merge(
        recommendations,
        on=["project_id", "indicator_id"],
        how="left",
        validate="one_to_one",
    )

    # Assessment Scores is the scoring source of truth. Repository quality is a fallback.
    source_quality = pd.to_numeric(merged["evidence_quality_score"], errors="coerce")
    repo_quality = pd.to_numeric(
        merged.get("repository_evidence_quality_score"), errors="coerce"
    )
    merged["evidence_quality_score"] = source_quality.fillna(repo_quality).fillna(0)

    defaults = {
        "evidence_count": 0,
        "evidence_title": "No linked evidence record",
        "evidence_type": "—",
        "evidence_source": "—",
        "source_organization": "—",
        "link_or_reference": "—",
        "evidence_notes": "No reviewer notes are available.",
        "recommended_action": "",
        "gap_type": "",
        "expected_benefit": "",
        "estimated_effort": "",
        "evidence_needed": "",
        "responsible_party_type": "",
    }
    for column, default in defaults.items():
        if column not in merged.columns:
            merged[column] = default
        merged[column] = merged[column].fillna(default)

    merged["why_it_applies"] = merged["why_it_applies"].fillna(
        merged["assessment_notes"]
    )
    merged["evidence_date_display"] = merged["evidence_date"].apply(
        lambda value: value.strftime("%Y-%m-%d") if pd.notna(value) else "—"
    )

    return enrich_assessments(merged)


@st.cache_data(show_spinner=False, ttl=30)
def load_all_data() -> dict[str, pd.DataFrame]:
    standards = load_standards()
    projects = load_projects()
    mapping = load_project_indicator_mapping()
    evidence = load_evidence()
    assessments = load_assessments()
    recommendations = load_recommendations()
    dashboard = merge_dashboard_data(
        assessments,
        standards,
        projects,
        mapping,
        evidence,
        recommendations,
    )
    return {
        "standards": standards,
        "projects": projects,
        "mapping": mapping,
        "evidence": evidence,
        "assessments": assessments,
        "recommendations": recommendations,
        "dashboard": dashboard,
    }


def filter_dashboard(
    dashboard: pd.DataFrame,
    domain: str | None = None,
    project_id: str | None = None,
    project_ids: Iterable[str] | None = None,
) -> pd.DataFrame:
    filtered = dashboard.copy()
    if domain and domain != "All domains":
        filtered = filtered[filtered["domain"] == domain]
    if project_id and project_id != "All projects":
        filtered = filtered[filtered["project_id"] == project_id]
    if project_ids:
        filtered = filtered[filtered["project_id"].isin(list(project_ids))]
    return filtered.reset_index(drop=True)
